"""
excel_exporter.py
-----------------
Xuất bản tin hàng ngày ra file Excel (.xlsx) để download từ Streamlit.

Dùng trong app.py:
    from excel_exporter import build_excel_bytes
    xlsx_bytes = build_excel_bytes(grouped)
    st.download_button("⬇️ Xuất Excel bản tin", data=xlsx_bytes, ...)
"""

from __future__ import annotations
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Cấu hình nhóm tin ──────────────────────────────────────────────────────
# key = tên nhóm trong grouped (group_by_category), value = màu header + nhãn
GROUP_CONFIG = {
    "THẾ GIỚI":    {"color": "C00000", "label": "THẾ GIỚI"},
    "TIN TỨC":     {"color": "1F497D", "label": "TRONG NƯỚC"},
    "DOANH NGHIỆP":{"color": "375623", "label": "DOANH NGHIỆP"},
}

# Thứ tự hiển thị trong Excel
GROUP_ORDER = ["THẾ GIỚI", "TIN TỨC", "DOANH NGHIỆP"]


def build_excel_bytes(grouped: dict) -> bytes:
    """
    Nhận dict {category: [NewsItem, ...]} từ group_by_category(),
    trả về bytes của file .xlsx sẵn sàng để st.download_button().
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    today = datetime.now().strftime("%d/%m/%Y")
    ws.title = f"Ban tin {today}"

    # Độ rộng cột
    ws.column_dimensions["A"].width = 3   # bullet / lề
    ws.column_dimensions["B"].width = 90  # tiêu đề tin

    thin = Side(style="thin", color="DDDDDD")
    border_bottom = Border(bottom=thin)

    row = 1

    for group_key in GROUP_ORDER:
        cfg = GROUP_CONFIG.get(group_key, {})
        color = cfg.get("color", "444444")
        label = cfg.get("label", group_key)
        items = grouped.get(group_key, [])

        # ── Header nhóm ──────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:B{row}")
        hcell = ws[f"A{row}"]
        hcell.value = label
        hcell.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        hcell.fill = PatternFill("solid", fgColor=color)
        hcell.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1, wrap_text=False)
        ws.row_dimensions[row].height = 24
        row += 1

        if not items:
            ws[f"B{row}"].value = "(Không có tin)"
            ws[f"B{row}"].font = Font(name="Arial", italic=True,
                                      color="888888", size=10)
            row += 2
            continue

        # ── Danh sách tin ────────────────────────────────────────────────
        for it in items:
            # Cột A: bullet
            bc = ws[f"A{row}"]
            bc.value = "•"
            bc.font = Font(name="Arial", size=10, color="555555")
            bc.alignment = Alignment(horizontal="center", vertical="top")

            # Cột B: tiêu đề (+ mã CK nếu có)
            ticker_tag = f"  [{'/'.join(it.tickers)}]" if it.tickers else ""
            tc = ws[f"B{row}"]
            tc.value = it.title + ticker_tag
            tc.font = Font(name="Arial", size=10, color="1A1A1A")
            tc.alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True)
            tc.border = border_bottom
            ws.row_dimensions[row].height = 28
            row += 1

        row += 1  # dòng trống giữa nhóm

    # ── Footer ───────────────────────────────────────────────────────────
    ws[f"B{row}"].value = (
        f"Nguồn: CafeF · Vietstock · VnExpress · VnEconomy · VietnamBiz  |  "
        f"Cập nhật: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws[f"B{row}"].font = Font(name="Arial", size=9, italic=True, color="AAAAAA")

    # Xuất ra bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
